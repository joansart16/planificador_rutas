from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.db.models import Count, Max, Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.utils.html import format_html
from django.urls import path, reverse
from django.utils.safestring import mark_safe
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Company, Contract, Driver, DriverUnavailability, Location, Route, ServiceTask, Vehicle


ROUTES_MENU_ORDER = {
    'Vehicle': 1,
    'Driver': 2,
    'Company': 3,
    'Contract': 4,
    'Route': 5,
    'ServiceTask': 6,
}

MODULE_OBRA   = 'OBRA'
MODULE_EVENTO = 'EVENTO'


def _current_module(request) -> str:
    """Returns the active module from the session, defaulting to OBRA."""
    return request.session.get('current_module', MODULE_OBRA)


def _user_allowed_modules(user) -> list[str]:
    if user.is_superuser:
        return [MODULE_OBRA, MODULE_EVENTO]
    if user.groups.filter(name='operadores_obra').exists():
        return [MODULE_OBRA]
    return [MODULE_OBRA]  # safe default


class ModuleFilterMixin:
    """Mixin that restricts querysets to the current session module."""

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        module = _current_module(request)
        allowed = _user_allowed_modules(request.user)
        if module not in allowed:
            module = allowed[0]
        return qs.filter(**{self._module_field: module})

    _module_field = 'module'  # override in subclasses if needed


_original_get_app_list = admin.AdminSite.get_app_list


def _get_app_list_with_custom_rutas_order(self, request, app_label=None):
    app_list = _original_get_app_list(self, request, app_label)
    for app in app_list:
        if app.get('app_label') == 'rutas':
            # Ocultar Location del menú: ahora está integrada dentro del formulario de Contrato
            app['models'] = [m for m in app['models'] if m.get('object_name') != 'Location']
            app['models'].sort(
                key=lambda model: (
                    ROUTES_MENU_ORDER.get(model.get('object_name'), 999),
                    model.get('name', ''),
                )
            )
    return app_list


admin.AdminSite.get_app_list = _get_app_list_with_custom_rutas_order

admin.site.site_header = 'Loorent · Planificador de Rutas'
admin.site.site_title  = 'Loorent'
admin.site.index_title = 'Panel de control'


class ExcelUploadForm(forms.Form):
    excel_file = forms.FileField(label='Archivo Excel (.xlsx)')


def _as_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _parse_bool(value, default=False) -> bool:
    if value is None or _as_text(value) == '':
        return default
    if isinstance(value, bool):
        return value
    normalized = _as_text(value).lower()
    return normalized in {'1', 'true', 'si', 'sí', 'yes', 'y', 'x'}


def _parse_choice(value, choices, default=None):
    if value is None or _as_text(value) == '':
        return default
    normalized = _as_text(value)
    by_key = {str(key): key for key, _ in choices}
    by_label = {str(label).lower(): key for key, label in choices}
    if normalized in by_key:
        return by_key[normalized]
    if normalized.lower() in by_label:
        return by_label[normalized.lower()]
    try:
        as_int = int(normalized)
        if str(as_int) in by_key:
            return by_key[str(as_int)]
    except (TypeError, ValueError):
        pass
    raise ValueError(f'Valor no válido: {normalized}')


def _parse_weekdays(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        items = value
    else:
        raw = _as_text(value)
        if not raw:
            return []
        items = [item.strip() for item in raw.split(',') if item.strip()]
    result = []
    for item in items:
        try:
            day = int(item)
        except (TypeError, ValueError) as exc:
            raise ValueError(f'Día inválido: {item}') from exc
        if day < 0 or day > 6:
            raise ValueError(f'Día fuera de rango (0-6): {item}')
        result.append(day)
    return sorted(set(result))


def _parse_date(value):
    if value is None or _as_text(value) == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _as_text(value)
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Fecha inválida: {raw}')


def _parse_time(value):
    if value is None or _as_text(value) == '':
        return None
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    raw = _as_text(value)
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(raw, fmt).time().replace(second=0, microsecond=0)
        except ValueError:
            continue
    raise ValueError(f'Hora inválida: {raw}')


def _parse_decimal(value):
    if value is None or _as_text(value) == '':
        return None
    raw = _as_text(value).replace(',', '.')
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f'Número decimal inválido: {raw}') from exc


def _parse_positive_int(value, default=None):
    if value is None or _as_text(value) == '':
        return default
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise ValueError(f'Número entero inválido: {value}') from exc
    if parsed <= 0:
        raise ValueError('El valor debe ser mayor que 0.')
    return parsed


class ExcelImportExportMixin:
    excel_template_columns = []
    excel_template_rows = []
    excel_instructions = []
    excel_value_guide = []
    excel_validations = {}
    excel_template_filename = ''

    def get_excel_template_filename(self) -> str:
        if self.excel_template_filename:
            return self.excel_template_filename
        return f'{self.model._meta.verbose_name_plural.lower().replace(" ", "_")}_plantilla.xlsx'

    def get_excel_upload_title(self) -> str:
        return f'Cargar Excel de {self.model._meta.verbose_name_plural}'

    def get_urls(self):
        urls = super().get_urls()
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        extra = [
            path(
                'cargar-excel/',
                self.admin_site.admin_view(self.upload_excel_view),
                name=f'{app_label}_{model_name}_upload_excel',
            ),
            path(
                'descargar-excel-prueba/',
                self.admin_site.admin_view(self.download_excel_template_view),
                name=f'{app_label}_{model_name}_download_excel_template',
            ),
        ]
        return extra + urls

    def _changelist_url(self):
        return reverse(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist')

    def download_excel_template_view(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Plantilla'
        sheet.append(self.excel_template_columns)
        for row in self.excel_template_rows:
            sheet.append(row)

        # Hoja alternativa con valores permitidos y significado.
        help_sheet = workbook.create_sheet('Valores')
        help_sheet.append(['Campo', 'Valores permitidos', 'Qué significa'])
        for row in self.excel_value_guide:
            help_sheet.append(row)

        # Listas desplegables en la hoja principal donde aplica.
        for field_name, options in self.excel_validations.items():
            if field_name not in self.excel_template_columns:
                continue
            if not options:
                continue
            col_index = self.excel_template_columns.index(field_name) + 1
            col_letter = get_column_letter(col_index)
            escaped = [str(option).replace('"', "'") for option in options]
            formula = '"' + ','.join(escaped) + '"'
            validation = DataValidation(type='list', formula1=formula, allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(f'{col_letter}2:{col_letter}500')

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_excel_template_filename()}"'
        return response

    def upload_excel_view(self, request):
        form = ExcelUploadForm(request.POST or None, request.FILES or None)
        if request.method == 'POST' and form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            try:
                workbook = load_workbook(filename=excel_file, data_only=True)
            except Exception:
                self.message_user(request, 'No se pudo leer el archivo Excel.', messages.ERROR)
                return redirect(self._changelist_url())

            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                self.message_user(request, 'El archivo está vacío.', messages.WARNING)
                return redirect(self._changelist_url())

            headers = [_as_text(col) for col in rows[0]]
            if headers != self.excel_template_columns:
                self.message_user(
                    request,
                    'Las columnas no coinciden con la plantilla. Descarga el Excel de prueba.',
                    messages.ERROR,
                )
                return redirect(self._changelist_url())

            created = 0
            skipped = 0
            warnings = []
            errors = []

            for row_number, values in enumerate(rows[1:], start=2):
                if not any(value is not None and _as_text(value) != '' for value in values):
                    continue
                row_data = {headers[idx]: values[idx] for idx in range(len(headers))}
                try:
                    import_result = self.import_excel_row(row_data)
                except Exception as exc:
                    errors.append(f'Fila {row_number}: {exc}')
                    continue

                result_status = import_result
                result_warning = ''
                if isinstance(import_result, tuple):
                    result_status = import_result[0]
                    result_warning = import_result[1] if len(import_result) > 1 else ''

                if result_status is True or result_status == 'created':
                    created += 1
                else:
                    skipped += 1
                    if not result_warning:
                        result_warning = 'Registro omitido (ya existía).'

                if result_warning:
                    warnings.append(f'Fila {row_number}: {result_warning}')

            if created or skipped:
                self.message_user(
                    request,
                    f'Importación completada: {created} creado(s), {skipped} omitido(s).',
                    messages.SUCCESS,
                )
            if errors:
                for err in errors[:10]:
                    self.message_user(request, err, messages.ERROR)
                if len(errors) > 10:
                    self.message_user(request, f'... y {len(errors) - 10} error(es) más.', messages.ERROR)
            if warnings:
                for warn in warnings[:10]:
                    self.message_user(request, warn, messages.WARNING)
                if len(warnings) > 10:
                    self.message_user(request, f'... y {len(warnings) - 10} aviso(s) más.', messages.WARNING)

            return redirect(self._changelist_url())

        return TemplateResponse(
            request,
            'admin/rutas/shared/upload_excel.html',
            {
                'title': self.get_excel_upload_title(),
                'form': form,
                'opts': self.model._meta,
                'changelist_url': self._changelist_url(),
                'template_url': reverse(
                    f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_download_excel_template'
                ),
                'excel_instructions': self.excel_instructions,
                'excel_template_columns': self.excel_template_columns,
            },
        )

    def import_excel_row(self, row_data) -> bool:
        raise NotImplementedError


class ContractAdminForm(forms.ModelForm):
    cleaning_weekdays = forms.TypedMultipleChoiceField(
        label='Dias de limpieza (lunes a domingo)',
        choices=Contract.Weekday.choices,
        coerce=int,
        widget=forms.CheckboxSelectMultiple,
        required=True,
        help_text='Selecciona exactamente los dias que correspondan a limpiezas por semana (1 a 7).',
    )

    class Meta:
        model = Contract
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['cleaning_weekdays'].initial = self.instance.cleaning_weekdays or []

    def clean_cleaning_weekdays(self):
        weekdays = self.cleaned_data.get('cleaning_weekdays') or []
        return sorted(set(int(day) for day in weekdays))


class ContractWithLocationForm(forms.ModelForm):
    """
    Formulario de Contrato con los campos de Ubicación integrados.
    Los nombres de los campos de Location coinciden con sus IDs de HTML (id_address,
    id_latitude, etc.) para que location_geocode.js funcione sin modificaciones.
    """

    # ── Campos de Ubicación ──────────────────────────────────────────────────
    name = forms.CharField(
        max_length=200,
        label='Nombre del sitio',
    )
    company = forms.ModelChoiceField(
        queryset=Company.objects.all(),
        required=False,
        label='Empresa / Cliente',
        help_text='Empresa titular. Permite filtrar todas sus ubicaciones.',
    )
    default_driver = forms.ModelChoiceField(
        queryset=Driver.objects.all(),
        required=False,
        label='Conductor por defecto',
        help_text='Se asignará automáticamente a las tareas de este sitio.',
    )
    contact_name = forms.CharField(
        max_length=150, required=False,
        label='Persona de contacto',
    )
    contact_phone = forms.CharField(
        max_length=20, required=False,
        label='Teléfono de contacto',
    )
    email = forms.EmailField(
        required=False,
        label='Email contacto obra',
        help_text='Opcional. Si se rellena, se usa en el exportador en lugar del email del cliente.',
    )
    comment = forms.CharField(
        widget=forms.Textarea(attrs={'rows': 3}),
        required=False,
        label='Comentario',
        help_text='Indicaciones operativas para el servicio en esta ubicación.',
    )
    cabin_count = forms.IntegerField(
        min_value=1,
        initial=1,
        label='Número de cabinas',
    )
    address = forms.CharField(
        widget=forms.Textarea(attrs={'rows': 2}),
        label='Dirección (calle y número)',
    )
    town = forms.CharField(
        max_length=100, required=False,
        label='Población',
        help_text='Núcleo habitado (ej: Palmanyola). Se rellena automáticamente.',
    )
    municipality = forms.CharField(
        max_length=100, required=False,
        label='Municipio',
        help_text='Municipio administrativo (ej: Bunyola). Se usa para calcular la comarca.',
    )
    postal_code = forms.CharField(
        max_length=10, required=False,
        label='Código postal',
    )
    zone = forms.ChoiceField(
        choices=[('', '---------')] + list(Location.Zone.choices),
        required=False,
        label='Zona de Mallorca',
        help_text='Comarca o zona de la isla.',
    )
    coords_cabin = forms.CharField(
        max_length=50, required=False,
        label='Coordenadas cabina',
        help_text='Pega desde Google Maps (ej: 39.619316, 2.643553).',
    )
    coords_entrance = forms.CharField(
        max_length=50, required=False,
        label='Coordenadas entrada finca',
        help_text='Opcional. Si se rellena, el exportador genera una fila extra con "Entrada finca".',
    )
    max_vehicle_size = forms.TypedChoiceField(
        choices=Vehicle.Size.choices,
        coerce=int,
        label='Tamaño máximo de vehículo',
        help_text='Vehículos con tamaño superior serán rechazados (gálibo).',
    )

    # ── Campos de Contrato ────────────────────────────────────────────────────
    module = forms.ChoiceField(
        choices=Contract.Module.choices,
        initial=Contract.Module.OBRA,
        label='Módulo',
        help_text='Obra o Evento.',
    )
    cleaning_weekdays = forms.TypedMultipleChoiceField(
        label='Días de limpieza (lunes a domingo)',
        choices=Contract.Weekday.choices,
        coerce=int,
        widget=forms.CheckboxSelectMultiple,
        required=True,
        help_text='Selecciona exactamente los días que correspondan a limpiezas por semana (1 a 7).',
    )

    class Meta:
        model = Contract
        exclude = ['location']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['cleaning_weekdays'].initial = self.instance.cleaning_weekdays or []

        # Pre-popular campos de ubicación al editar un contrato existente
        if self.instance.pk and self.instance.location_id:
            loc = self.instance.location
            self.initial.update({
                'name': loc.name,
                'company': loc.company_id,
                'default_driver': loc.default_driver_id,
                'contact_name': loc.contact_name,
                'contact_phone': loc.contact_phone,
                'email': loc.email,
                'comment': loc.comment,
                'cabin_count': loc.cabin_count,
                'address': loc.address,
                'town': loc.town,
                'municipality': loc.municipality,
                'postal_code': loc.postal_code,
                'zone': loc.zone,
                'coords_cabin': loc.coords_cabin,
                'coords_entrance': loc.coords_entrance,
                'max_vehicle_size': loc.max_vehicle_size,
            })

    def clean_cleaning_weekdays(self):
        weekdays = self.cleaned_data.get('cleaning_weekdays') or []
        return sorted(set(int(day) for day in weekdays))


class DriverAdminForm(forms.ModelForm):
    working_days = forms.TypedMultipleChoiceField(
        label='Días de trabajo',
        choices=Driver.Weekday.choices,
        coerce=int,
        widget=forms.CheckboxSelectMultiple,
        required=False,
        help_text='Selecciona los días de lunes a domingo en los que trabaja.',
    )

    class Meta:
        model = Driver
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['working_days'].initial = self.instance.working_days or []

    def clean_working_days(self):
        days = self.cleaned_data.get('working_days') or []
        return sorted(set(int(day) for day in days))


class LocationAdminForm(forms.ModelForm):
    class Meta:
        model = Location
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['contact_name'].label = 'Pers. contacto'
        self.fields['contact_name'].help_text = 'Persona de contacto'
        self.fields['contact_phone'].label = 'Tel. contacto'
        self.fields['contact_phone'].help_text = 'Teléfono de contacto'
        self.fields['town'].label = 'Población'
        self.fields['town'].help_text = 'Población'
        self.fields['cabin_count'].label = 'N. cab.'
        self.fields['cabin_count'].help_text = 'Número cabinas'
        self.fields['default_driver'].label = 'Cond. def.'
        self.fields['default_driver'].help_text = 'Conductor por defecto'


class DriverAvailabilityFilter(admin.SimpleListFilter):
    title = 'Disponibilidad'
    parameter_name = 'availability'

    def lookups(self, request, model_admin):
        return [
            ('available', 'Disponible mañana'),
            ('unavailable', 'No disponible mañana'),
            ('non_working', 'No laboral mañana'),
        ]

    def queryset(self, request, queryset):
        from datetime import timedelta

        target_date = timezone.localdate() + timedelta(days=1)
        weekday = target_date.weekday()
        unavailable_driver_ids = DriverUnavailability.objects.filter(
            start_date__lte=target_date,
            end_date__gte=target_date,
        ).values_list('driver_id', flat=True)

        if self.value() == 'available':
            return queryset.filter(working_days__contains=weekday).exclude(id__in=unavailable_driver_ids)
        if self.value() == 'unavailable':
            return queryset.filter(id__in=unavailable_driver_ids)
        if self.value() == 'non_working':
            return queryset.exclude(working_days__contains=weekday)
        return queryset


# ──────────────────────────────────────────────────────────────────────────────
# Company
# ──────────────────────────────────────────────────────────────────────────────
@admin.register(Company)
class CompanyAdmin(ExcelImportExportMixin, admin.ModelAdmin):
    change_list_template = 'admin/rutas/company/change_list.html'
    excel_template_filename = 'empresas_plantilla.xlsx'
    excel_template_columns = ['razon_social', 'correo_electronico']
    excel_template_rows = [['Empresa Demo', 'admin@empresa-demo.com']]
    excel_value_guide = [
        ('razon_social', 'Texto libre', 'Nombre legal de la empresa.'),
        ('correo_electronico', 'Correo válido o vacío', 'Email de contacto principal.'),
    ]

    list_display  = ('name', 'email', 'location_count')
    search_fields = ('name', 'email')
    ordering      = ('name',)
    actions       = ['delete_selected']
    fieldsets     = (
        (None, {
            'fields': ('name', 'email'),
        }),
    )

    def get_queryset(self, request):
        return super().get_queryset(request).annotate(location_count=Count('locations'))

    @admin.display(description='Ubicaciones', ordering='location_count')
    def location_count(self, obj: Company) -> int:
        return obj.location_count

    def import_excel_row(self, row_data) -> bool:
        name = _as_text(row_data['razon_social'])
        email = _as_text(row_data['correo_electronico'])
        if not name:
            raise ValueError('razon_social es obligatorio.')
        if Company.objects.filter(name=name).exists():
            return ('skipped', f'Empresa omitida: "{name}" ya existe.')
        Company.objects.create(name=name, email=email)
        return 'created'


# ──────────────────────────────────────────────────────────────────────────────
# Driver
# ──────────────────────────────────────────────────────────────────────────────
# Inline: indisponibilidades del conductor
class DriverUnavailabilityInline(admin.TabularInline):
    model            = DriverUnavailability
    extra            = 1
    fields           = ('reason', 'start_date', 'end_date', 'notes')
    ordering         = ['-start_date']


@admin.register(Driver)
class DriverAdmin(ExcelImportExportMixin, admin.ModelAdmin):
    change_list_template = 'admin/rutas/driver/change_list.html'
    excel_template_filename = 'conductores_plantilla.xlsx'
    excel_template_columns = ['nombre', 'dias_trabajo']
    excel_template_rows = [['Juan Pérez', '0,1,2,3,4,5']]
    excel_instructions = [
        'dias_trabajo: números separados por coma entre 0 y 6 (0=Lunes, 6=Domingo).',
    ]
    excel_value_guide = [
        ('nombre', 'Texto libre', 'Nombre completo del conductor.'),
        ('dias_trabajo', '0..6 separados por coma', '0=Lunes, 1=Martes, 2=Miércoles, 3=Jueves, 4=Viernes, 5=Sábado, 6=Domingo.'),
    ]

    form = DriverAdminForm
    list_display  = ('name', 'working_days_badge', 'current_availability_badge')
    list_filter   = (DriverAvailabilityFilter,)
    search_fields = ('name',)
    ordering      = ('name',)
    actions       = ['delete_selected']
    inlines       = [DriverUnavailabilityInline]
    fieldsets     = (
        ('Información personal', {
            'fields': ('name',),
        }),
        ('Horario laboral', {
            'fields': ('working_days',),
            'description': 'Selecciona los días de la semana en los que trabaja.',
        }),
    )

    @admin.display(description='Días de trabajo')
    def working_days_badge(self, obj: Driver) -> str:
        if not obj.working_days:
            return '—'
        day_names = dict(Driver.Weekday.choices)
        selected = [day_names.get(int(day), str(day)) for day in obj.working_days]
        return ', '.join(selected)

    @admin.display(description='Disponibilidad mañana')
    def current_availability_badge(self, obj: Driver) -> str:
        from datetime import timedelta

        target_date = timezone.localdate() + timedelta(days=1)
        if not obj.is_working_day(target_date.weekday()):
            return 'No laboral'
        unavailability = obj.get_active_unavailability(target_date)
        if unavailability:
            return f'No disponible ({unavailability.get_reason_display()})'
        return 'Disponible'

    def import_excel_row(self, row_data) -> bool:
        name = _as_text(row_data['nombre'])
        working_days = _parse_weekdays(row_data['dias_trabajo'])
        if not name:
            raise ValueError('nombre es obligatorio.')
        if Driver.objects.filter(name=name).exists():
            return ('skipped', f'Conductor omitido: "{name}" ya existe.')
        Driver.objects.create(name=name, working_days=working_days)
        return 'created'


# ──────────────────────────────────────────────────────────────────────────────
# Vehicle
# ──────────────────────────────────────────────────────────────────────────────
@admin.register(Vehicle)
class VehicleAdmin(ExcelImportExportMixin, admin.ModelAdmin):
    change_list_template = 'admin/rutas/vehicle/change_list.html'
    excel_template_filename = 'vehiculos_plantilla.xlsx'
    excel_template_columns = ['nombre_vehiculo', 'matricula', 'tipo', 'estado']
    excel_template_rows = [['Pickup 1', '1234ABC', 'Camión grande', 'Disponible']]
    excel_instructions = [
        'tipo permitido: Pickup, Camión pequeño, Camión grande.',
        'estado permitido: Disponible, En mantenimiento, Retirado.',
        'Las tareas de entrega y recogida solo admiten tamaño 3 (Camión grande).',
    ]
    excel_value_guide = [
        ('nombre_vehiculo', 'Texto libre o vacío', 'Nombre interno del vehículo.'),
        ('matricula', 'Texto único', 'Identificador único del vehículo.'),
        ('tipo', 'Pickup, Camión pequeño, Camión grande', 'Tipo del vehículo.'),
        ('estado', 'Disponible, En mantenimiento, Retirado', 'Estado operativo del vehículo.'),
    ]
    excel_validations = {
        'tipo': ['Pickup', 'Camión pequeño', 'Camión grande'],
        'estado': ['Disponible', 'En mantenimiento', 'Retirado'],
    }

    list_display  = ('name', 'license_plate', 'get_size_display_label', 'status')
    list_filter   = ('status', 'size')
    search_fields = ('name', 'license_plate')
    ordering      = ('license_plate',)
    actions       = ['delete_selected']

    @admin.display(description='Tamaño', ordering='size')
    def get_size_display_label(self, obj: Vehicle) -> str:
        return obj.get_size_display()

    def import_excel_row(self, row_data) -> bool:
        name = _as_text(row_data['nombre_vehiculo'])
        license_plate = _as_text(row_data['matricula']).upper()
        if not license_plate:
            raise ValueError('matricula es obligatorio.')

        if Vehicle.objects.filter(license_plate=license_plate).exists():
            return ('skipped', f'Vehículo omitido: matrícula "{license_plate}" ya existe.')

        size = _parse_choice(row_data['tipo'], Vehicle.Size.choices)
        status = _parse_choice(row_data['estado'], Vehicle.Status.choices, default=Vehicle.Status.AVAILABLE)
        Vehicle.objects.create(
            name=name,
            license_plate=license_plate,
            size=size,
            status=status,
        )
        return 'created'


# ──────────────────────────────────────────────────────────────────────────────
# Location
# ──────────────────────────────────────────────────────────────────────────────
@admin.register(Location)
class LocationAdmin(ExcelImportExportMixin, admin.ModelAdmin):
    form = LocationAdminForm
    change_list_template = 'admin/rutas/location/change_list.html'
    excel_template_filename = 'ubicaciones_plantilla.xlsx'
    excel_template_columns = [
        'nombre', 'empresa', 'contacto_nombre', 'contacto_telefono', 'email_obra',
        'direccion', 'comentario', 'poblacion', 'municipio', 'codigo_postal', 'zona',
        'cabinas', 'tipo_max_vehiculo', 'conductor_por_defecto',
        'coords_cabina', 'coords_entrada',
    ]
    excel_template_rows = [[
        'Obra Son Vida', 'Empresa Demo', 'Miguel', '600123123', 'obra@demo.com',
        'Calle Falsa 123', 'Ir con pickup o mpw',
        'Palma', 'Palma', '07010', 'PALMA', 1, 'Camión pequeño', 'Juan Pérez',
        '39.595000, 2.650000', '',
    ]]
    excel_instructions = [
        'tipo_max_vehiculo permitido: Pickup, Camión pequeño, Camión grande.',
        'zona permitida: PALMA, TRAMUNTANA, RAIGUER, PLA, MIGJORN, LLEVANT (opcional).',
        'empresa y conductor_por_defecto deben existir previamente o dejarse en blanco.',
        'coords_cabina y coords_entrada: formato "lat, lon" (ej: 39.619316, 2.643553).',
        'email_obra: opcional, sobreescribe el email del cliente en el exportador.',
    ]
    excel_value_guide = [
        ('nombre', 'Texto único', 'Nombre de la ubicación.'),
        ('empresa', 'Razón social existente o vacío', 'Empresa titular de la ubicación.'),
        ('contacto_nombre', 'Texto libre', 'Persona de contacto en obra.'),
        ('contacto_telefono', 'Texto libre', 'Teléfono de contacto.'),
        ('email_obra', 'Email válido o vacío', 'Email de la ubicación. Si se rellena, se usa en lugar del email del cliente.'),
        ('direccion', 'Texto libre', 'Calle y número.'),
        ('comentario', 'Texto libre', 'Comentario operativo de la ubicación.'),
        ('poblacion', 'Texto libre', 'Población.'),
        ('municipio', 'Texto libre', 'Municipio administrativo.'),
        ('codigo_postal', 'Texto o número', 'Código postal de la ubicación.'),
        ('zona', 'PALMA/TRAMUNTANA/RAIGUER/PLA/MIGJORN/LLEVANT o vacío', 'Zona para filtrado logístico.'),
        ('cabinas', 'Entero > 0', 'Número de cabinas de la ubicación.'),
        ('tipo_max_vehiculo', 'Pickup, Camión pequeño, Camión grande', 'Tipo máximo permitido en la ubicación.'),
        ('conductor_por_defecto', 'Nombre de conductor existente o vacío', 'Conductor habitual de la ubicación.'),
        ('coords_cabina', '"lat, lon" o vacío', 'Coordenadas de la cabina (ej: 39.619316, 2.643553).'),
        ('coords_entrada', '"lat, lon" o vacío', 'Coordenadas entrada finca. Si existe, genera fila extra en exportación.'),
    ]
    excel_validations = {
        'cabinas': ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10'],
        'tipo_max_vehiculo': ['Pickup', 'Camión pequeño', 'Camión grande'],
        'zona': ['PALMA', 'TRAMUNTANA', 'RAIGUER', 'PLA', 'MIGJORN', 'LLEVANT'],
    }

    list_display   = (
        'name', 'company', 'default_driver_short', 'cabin_count_short',
        'contact_name_short', 'contact_phone_short',
        'address_short', 'coords_cabin_short', 'town_short', 'postal_code',
        'get_zone_label', 'get_max_size_label',
    )
    list_filter    = ('zone', 'max_vehicle_size', 'company')
    search_fields  = (
        'name', 'company__name', 'address', 'town', 'municipality', 'postal_code',
        'contact_name', 'contact_phone',
    )
    ordering       = ('company__name', 'name')
    actions        = ['delete_selected']
    autocomplete_fields = ('company', 'default_driver')
    fieldsets      = (
        ('Identificación', {
            'fields': ('name', 'company', 'default_driver'),
        }),
        ('Contacto en obra', {
            'fields': ('contact_name', 'contact_phone', 'email', 'comment', 'cabin_count'),
            'description': 'Responsable de la ubicación.',
        }),
        ('Dirección y coordenadas', {
            'fields': ('address', 'town', 'municipality', 'postal_code', 'zone', 'coords_cabin', 'coords_entrance'),
        }),
        ('Restricción de vehículo', {
            'fields': ('max_vehicle_size',),
        }),
    )

    @admin.display(description='Zona', ordering='zone')
    def get_zone_label(self, obj: Location) -> str:
        return obj.get_zone_display() if obj.zone else '—'

    @admin.display(description=mark_safe('<span title="Conductor por defecto">Cond. def.</span>'), ordering='default_driver')
    def default_driver_short(self, obj: Location) -> str:
        return str(obj.default_driver) if obj.default_driver else '—'

    @admin.display(description=mark_safe('<span title="Número cabinas">N. cab.</span>'), ordering='cabin_count')
    def cabin_count_short(self, obj: Location) -> str:
        return format_html('<div style="text-align:center;">{}</div>', obj.cabin_count)

    @admin.display(description=mark_safe('<span title="Persona de contacto">Pers. contacto</span>'), ordering='contact_name')
    def contact_name_short(self, obj: Location) -> str:
        return obj.contact_name or '—'

    @admin.display(description=mark_safe('<span title="Teléfono de contacto">Tel. contacto</span>'), ordering='contact_phone')
    def contact_phone_short(self, obj: Location) -> str:
        return obj.contact_phone or '—'

    @admin.display(description=mark_safe('<span title="Dirección">Dir.</span>'), ordering='address')
    def address_short(self, obj: Location) -> str:
        return obj.address or '—'

    @admin.display(description=mark_safe('<span title="Coordenadas cabina">Coords.</span>'), ordering='coords_cabin')
    def coords_cabin_short(self, obj: Location) -> str:
        return obj.coords_cabin or '—'

    @admin.display(description=mark_safe('<span title="Población">Población</span>'), ordering='town')
    def town_short(self, obj: Location) -> str:
        return obj.town or '—'

    @admin.display(description='Tam. máx. vehículo', ordering='max_vehicle_size')
    def get_max_size_label(self, obj: Location) -> str:
        return obj.get_max_vehicle_size_display()

    def import_excel_row(self, row_data) -> bool:
        name = _as_text(row_data['nombre'])
        if not name:
            raise ValueError('nombre es obligatorio.')

        if Location.objects.filter(name=name).exists():
            return ('skipped', f'Ubicación omitida: "{name}" ya existe.')

        company = None
        company_name = _as_text(row_data['empresa'])
        if company_name:
            company = Company.objects.filter(name=company_name).first()
            if not company:
                raise ValueError(f'No existe empresa={company_name}')

        default_driver = None
        driver_name = _as_text(row_data['conductor_por_defecto'])
        if driver_name:
            default_driver = Driver.objects.filter(name=driver_name).first()
            if not default_driver:
                raise ValueError(f'No existe conductor={driver_name}')

        max_vehicle_size = _parse_choice(row_data['tipo_max_vehiculo'], Vehicle.Size.choices)
        zone = _as_text(row_data['zona'])
        if zone:
            zone = _parse_choice(zone, Location.Zone.choices)

        data = {
            'company': company,
            'contact_name': _as_text(row_data['contacto_nombre']),
            'contact_phone': _as_text(row_data['contacto_telefono']),
            'address': _as_text(row_data['direccion']),
            'comment': _as_text(row_data['comentario']),
            'town': _as_text(row_data['poblacion']),
            'municipality': _as_text(row_data['municipio']),
            'postal_code': _as_text(row_data['codigo_postal']),
            'zone': zone or '',
            'cabin_count': _parse_positive_int(row_data['cabinas'], default=1),
            'email': _as_text(row_data['email_obra']),
            'max_vehicle_size': max_vehicle_size,
            'default_driver': default_driver,
            'coords_cabin': _as_text(row_data['coords_cabina']),
            'coords_entrance': _as_text(row_data['coords_entrada']),
        }
        if not data['address']:
            raise ValueError('direccion es obligatorio.')

        Location.objects.create(
            name=name,
            **data,
        )
        return 'created'


# ──────────────────────────────────────────────────────────────────────────────
# Inline: tareas del contrato (solo lectura de tipo/fecha; asignación de chofer/vehículo)
# ──────────────────────────────────────────────────────────────────────────────
class ServiceTaskInline(admin.TabularInline):
    model            = ServiceTask
    extra            = 0
    fields           = ('task_type', 'scheduled_date', 'suggested_vehicle_size', 'driver', 'vehicle')
    readonly_fields  = ('task_type', 'scheduled_date', 'suggested_vehicle_size')
    ordering         = ('scheduled_date',)
    show_change_link = True


# ──────────────────────────────────────────────────────────────────────────────
# Contract
# ──────────────────────────────────────────────────────────────────────────────
@admin.register(Contract)
class ContractAdmin(ModuleFilterMixin, ExcelImportExportMixin, admin.ModelAdmin):
    _module_field = 'module'
    excel_template_filename = 'contratos_plantilla.xlsx'
    excel_template_columns = [
        'ubicacion', 'fecha_inicio', 'fecha_fin', 'limpiezas_semana', 'dias_limpieza',
        'hora_acceso_inicio', 'hora_acceso_fin', 'estado',
    ]
    excel_template_rows = [[
        'Obra Son Vida', '2026-05-06', '', 3, '0,2,4', '08:00', '18:00', 'Activo',
    ]]
    excel_instructions = [
        'dias_limpieza: números separados por coma entre 0 y 6 (0=Lunes, 6=Domingo).',
        'estado permitido: ACTIVE o CLOSED (o su etiqueta en español).',
        'Si el contrato ya existe (misma ubicación + fecha_inicio), se omitirá.',
    ]
    excel_value_guide = [
        ('ubicacion', 'Nombre de ubicación existente', 'Ubicación a la que pertenece el contrato.'),
        ('fecha_inicio', 'YYYY-MM-DD o DD/MM/YYYY', 'Fecha de alta del contrato.'),
        ('fecha_fin', 'YYYY-MM-DD o DD/MM/YYYY o vacío', 'Fecha de recogida/fin.'),
        ('limpiezas_semana', '1..7', 'Número de limpiezas semanales.'),
        ('dias_limpieza', '0..6 separados por coma', '0=Lunes, 1=Martes, 2=Miércoles, 3=Jueves, 4=Viernes, 5=Sábado, 6=Domingo.'),
        ('hora_acceso_inicio', 'HH:MM o vacío', 'Hora mínima de acceso.'),
        ('hora_acceso_fin', 'HH:MM o vacío', 'Hora máxima de acceso.'),
        ('estado', 'Activo, Interrumpido o Retirado', 'Estado administrativo del pedido.'),
    ]
    excel_validations = {
        'limpiezas_semana': ['1', '2', '3', '4', '5', '6', '7'],
        'estado': ['Activo', 'Interrumpido', 'Retirado'],
    }

    form = ContractWithLocationForm
    change_list_template = 'admin/rutas/contract/change_list.html'
    list_display   = (
        'budget_number', '__str__', 'location', 'start_date', 'end_date',
        'cleaning_frequency', 'cleaning_days_badge', 'coherence_warning_badge', 'status',
    )
    list_filter    = ('status', 'location__zone', 'location__company', 'start_date')
    search_fields  = ('location__name', 'location__company__name', 'location__town', 'budget_number')
    date_hierarchy = 'start_date'
    inlines        = [ServiceTaskInline]
    actions        = ['delete_selected']
    fieldsets      = (
        ('Módulo', {
            'fields': ('module',),
            'description': 'Indica si este pedido pertenece a Obra o a Evento.',
        }),
        ('Presupuesto', {
            'fields': ('budget_number',),
        }),
        ('Identificación de la ubicación', {
            'fields': ('name', 'company', 'default_driver'),
        }),
        ('Contacto en obra', {
            'fields': ('contact_name', 'contact_phone', 'email', 'comment', 'cabin_count'),
            'description': 'Responsable de la ubicación.',
        }),
        ('Dirección y coordenadas', {
            'fields': ('address', 'town', 'municipality', 'postal_code', 'zone', 'coords_cabin', 'coords_entrance'),
            'description': (
                'Dirección textual de la ubicación. '
                'Coordenadas: pega directamente desde Google Maps (ej: 39.619316, 2.643553).'
            ),
        }),
        ('Restricción de vehículo', {
            'fields': ('max_vehicle_size',),
        }),
        ('Período del contrato', {
            'fields': ('start_date', 'end_date', 'cleaning_frequency', 'cleaning_weekdays'),
        }),
        ('Restricciones horarias de acceso', {
            'fields': ('access_start_time', 'access_end_time'),
            'classes': ('collapse',),
            'description': (
                'Opcional. Ventana horaria en la que se permite el acceso al sitio. '
                'Solo informativo para los administrativos.'
            ),
        }),
        ('Estado', {
            'fields': ('status',),
        }),
    )

    class Media:
        js = ('rutas/admin/contract_dates_tomorrow.js',)

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['google_maps_api_key'] = settings.GOOGLE_MAPS_API_KEY
        return super().changeform_view(request, object_id, form_url, extra_context)

    def save_model(self, request, obj, form, change):
        # Auto-set module from session when creating a new contract
        if not change and not obj.module:
            obj.module = _current_module(request)
        cd = form.cleaned_data
        location_data = {
            'name': cd['name'],
            'company': cd.get('company'),
            'default_driver': cd.get('default_driver'),
            'contact_name': cd.get('contact_name', ''),
            'contact_phone': cd.get('contact_phone', ''),
            'email': cd.get('email', ''),
            'comment': cd.get('comment', ''),
            'cabin_count': cd.get('cabin_count', 1),
            'address': cd['address'],
            'town': cd.get('town', ''),
            'municipality': cd.get('municipality', ''),
            'postal_code': cd.get('postal_code', ''),
            'zone': cd.get('zone', ''),
            'coords_cabin': cd.get('coords_cabin', ''),
            'coords_entrance': cd.get('coords_entrance', ''),
            'max_vehicle_size': cd['max_vehicle_size'],
        }
        if obj.location_id:
            loc = obj.location
            for attr, value in location_data.items():
                setattr(loc, attr, value)
            loc.save()
        else:
            loc = Location(**location_data)
            loc.save()
            obj.location = loc
        super().save_model(request, obj, form, change)

    def import_excel_row(self, row_data) -> bool:
        location_name = _as_text(row_data['ubicacion'])
        if not location_name:
            raise ValueError('ubicacion es obligatorio.')
        location = Location.objects.filter(name=location_name).first()
        if not location:
            raise ValueError(f'No existe ubicacion={location_name}')

        start_date = _parse_date(row_data['fecha_inicio'])
        if not start_date:
            raise ValueError('fecha_inicio es obligatorio.')

        if Contract.objects.filter(location=location, start_date=start_date).exists():
            return (
                'skipped',
                f'Contrato omitido: ya existe para ubicación "{location_name}" con fecha_inicio "{start_date}".',
            )

        end_date = _parse_date(row_data['fecha_fin'])
        cleaning_weekdays = _parse_weekdays(row_data['dias_limpieza'])
        cleaning_frequency = int(row_data['limpiezas_semana']) if row_data['limpiezas_semana'] is not None else 0
        status = _parse_choice(row_data['estado'], Contract.Status.choices, default=Contract.Status.ACTIVE)

        if not cleaning_frequency:
            raise ValueError('limpiezas_semana es obligatorio.')
        is_coherent = cleaning_frequency == len(cleaning_weekdays)

        obj = Contract.objects.create(
            location=location,
            start_date=start_date,
            end_date=end_date,
            cleaning_frequency=cleaning_frequency,
            cleaning_weekdays=cleaning_weekdays,
            access_start_time=_parse_time(row_data['hora_acceso_inicio']),
            access_end_time=_parse_time(row_data['hora_acceso_fin']),
            status=status,
        )
        if is_coherent:
            obj.full_clean()
            obj.save()
            return 'created'
        return (
            'created',
            (
                f'Contrato #{obj.pk} importado con incoherencia: '
                f'limpiezas_semana={cleaning_frequency} pero dias_limpieza={len(cleaning_weekdays)}.'
            ),
        )

    @admin.display(description='Coherencia')
    def coherence_warning_badge(self, obj: Contract) -> str:
        count_days = len(obj.cleaning_weekdays or [])
        if obj.cleaning_frequency == count_days:
            return 'OK'
        edit_url = reverse('admin:rutas_contract_change', args=[obj.pk])
        return format_html(
            'Este contrato no es coherente. <a href="{}">Rectificar</a>',
            edit_url,
        )

    @admin.display(description='Dias limpieza')
    def cleaning_days_badge(self, obj: Contract) -> str:
        day_names = dict(Contract.Weekday.choices)
        selected = [day_names.get(int(day), str(day)) for day in (obj.cleaning_weekdays or [])]
        return ', '.join(selected) if selected else '—'

    def delete_model(self, request, obj):
        """Mostrar advertencia de tareas que se van a eliminar en cascada."""
        task_count = obj.tasks.count()
        if task_count > 0:
            msg = (
                f'Se eliminará el contrato "{obj}" junto con '
                f'{task_count} tarea{"s" if task_count > 1 else ""} asociada{"s" if task_count > 1 else ""}.'
            )
            self.message_user(request, msg, messages.WARNING)
        super().delete_model(request, obj)

    # ------------------------------------------------------------------
    # URL extra: vista de generación manual de limpiezas por día
    # ------------------------------------------------------------------
    def get_urls(self):
        urls = super().get_urls()
        extra = [
            path(
                'generar-tareas/',
                self.admin_site.admin_view(self.generate_tasks_view),
                name='rutas_contract_generate_tasks',
            ),
        ]
        return extra + urls

    def generate_tasks_view(self, request):
        """Vista para generar manualmente las LIMPIEZAS de un día concreto."""
        from datetime import date, timedelta
        from .models import ServiceTask

        today = timezone.localdate()
        tomorrow = today + timedelta(days=1)

        # Fechas rápidas para los botones
        quick_dates = {
            'Mañana':        tomorrow,
            'Pasado mañana': today + timedelta(days=2),
            'En 3 días':     today + timedelta(days=3),
        }

        generated = None
        skipped   = None
        target_date = None

        if request.method == 'POST':
            date_str = request.POST.get('target_date', '').strip()
            try:
                target_date = date.fromisoformat(date_str)
            except ValueError:
                self.message_user(request, 'Fecha no válida.', messages.ERROR)
            else:
                weekday = target_date.weekday()
                contracts = Contract.objects.filter(
                    status=Contract.Status.ACTIVE,
                    module=_current_module(request),
                    start_date__lte=target_date,
                    cleaning_weekdays__contains=weekday,
                ).filter(
                    Q(end_date__isnull=True) | Q(end_date__gte=target_date)
                ).select_related('location')

                new_tasks = []
                skipped_list = []
                for contract in contracts:
                    already = ServiceTask.objects.filter(
                        contract=contract,
                        task_type=ServiceTask.TaskType.LIMPIEZA,
                        scheduled_date=target_date,
                    ).exists()
                    blocked_by_delivery = ServiceTask.objects.filter(
                        contract=contract,
                        task_type__in=[
                            ServiceTask.TaskType.ENTREGA,
                            ServiceTask.TaskType.RECOGIDA,
                        ],
                        scheduled_date=target_date,
                    ).exists()
                    if already or blocked_by_delivery:
                        skipped_list.append((contract, 'ya existe' if already else 'hay entrega/recogida ese día'))
                    else:
                        new_tasks.append(ServiceTask(
                            task_type=ServiceTask.TaskType.LIMPIEZA,
                            scheduled_date=target_date,
                            location=contract.location,
                            contract=contract,
                            driver=None,
                            vehicle=None,
                            suggested_vehicle_size=contract.location.max_vehicle_size,
                        ))

                ServiceTask.objects.bulk_create(new_tasks)
                generated = new_tasks
                skipped   = skipped_list

                if new_tasks:
                    self.message_user(
                        request,
                        f'Generadas {len(new_tasks)} tarea(s) de limpieza para el {target_date}.',
                        messages.SUCCESS,
                    )
                else:
                    self.message_user(
                        request,
                        f'No hay contratos activos con limpieza ese día ({target_date}) '
                        'o las tareas ya existían.',
                        messages.WARNING,
                    )

        return TemplateResponse(
            request,
            'admin/rutas/contract/generate_tasks.html',
            {
                'title':        'Generar limpiezas por día',
                'quick_dates':  quick_dates,
                'today':        today,
                'tomorrow':     tomorrow,
                'target_date':  target_date,
                'generated':    generated,
                'skipped':      skipped,
                'opts':         self.model._meta,
            },
        )


# ──────────────────────────────────────────────────────────────────────────────
# Route
# ──────────────────────────────────────────────────────────────────────────────
import json as _json


class RouteStopInline(admin.TabularInline):
    model   = ServiceTask
    fk_name = 'route'
    extra   = 0
    fields  = ('route_order', 'task_type', 'scheduled_date', 'location', 'driver', 'vehicle')
    readonly_fields = ('task_type', 'scheduled_date', 'location')
    ordering = ('route_order',)
    verbose_name        = 'Parada'
    verbose_name_plural = 'Paradas asignadas'

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('location', 'driver', 'vehicle')


@admin.register(Route)
class RouteAdmin(ModuleFilterMixin, admin.ModelAdmin):
    _module_field = 'module'

    list_display   = ('date', 'module', 'driver', 'vehicle', 'name', 'stop_count', 'map_link')
    list_filter    = ('date', 'module', 'driver', 'vehicle')
    search_fields  = ('driver__name', 'vehicle__license_plate', 'name')
    date_hierarchy = 'date'
    inlines        = [RouteStopInline]
    autocomplete_fields = ('driver', 'vehicle')
    fieldsets = (
        (None, {'fields': ('date', 'module', 'driver', 'vehicle', 'name')}),
    )

    def get_urls(self):
        urls = super().get_urls()
        extra = [
            path(
                '<int:pk>/mapa/',
                self.admin_site.admin_view(self.mapa_view),
                name='rutas_route_mapa',
            ),
            path(
                '<int:pk>/reorder/',
                self.admin_site.admin_view(self.reorder_view),
                name='rutas_route_reorder',
            ),
            path(
                '<int:pk>/add-stop/',
                self.admin_site.admin_view(self.add_stop_view),
                name='rutas_route_add_stop',
            ),
        ]
        return extra + urls

    @admin.display(description='Paradas', ordering='stop_count')
    def stop_count(self, obj: Route) -> int:
        return obj.stop_count

    @admin.display(description='Mapa')
    def map_link(self, obj: Route) -> str:
        url = reverse('admin:rutas_route_mapa', args=[obj.pk])
        return format_html('<a href="{}" target="_blank">🗺️ Ver mapa</a>', url)

    def get_queryset(self, request):
        return super().get_queryset(request).annotate(stop_count=Count('stops'))

    def save_model(self, request, obj, form, change):
        if not change and not obj.module:
            obj.module = _current_module(request)
        super().save_model(request, obj, form, change)

    def mapa_view(self, request, pk: int):
        route = Route.objects.select_related('driver', 'vehicle').get(pk=pk)

        stops_qs = (
            ServiceTask.objects.filter(route=route)
            .select_related('location__company', 'contract')
            .order_by('route_order', 'pk')
        )

        stops = []
        for task in stops_qs:
            lat = lng = ''
            has_coords = False
            if task.location and task.location.coords_cabin:
                try:
                    parts = task.location.coords_cabin.split(',')
                    lat, lng = parts[0].strip(), parts[1].strip()
                    float(lat); float(lng)
                    has_coords = True
                except (ValueError, IndexError):
                    pass
            stops.append({'task': task, 'lat': lat, 'lng': lng, 'has_coords': has_coords})

        # Unassigned tasks for the same date + module
        unassigned = (
            ServiceTask.objects.filter(
                scheduled_date=route.date,
                contract__module=route.module,
                route__isnull=True,
            )
            .select_related('location', 'contract')
            .order_by('location__name')
        )

        return TemplateResponse(
            request,
            'admin/rutas/route/mapa.html',
            {
                'route': route,
                'stops': stops,
                'unassigned': unassigned,
                'date': route.date,
                'google_maps_api_key': settings.GOOGLE_MAPS_API_KEY,
                'opts': Route._meta,
                'csrf_token': request.META.get('CSRF_COOKIE', ''),
            },
        )

    def reorder_view(self, request, pk: int):
        if request.method != 'POST':
            from django.http import JsonResponse
            return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
        from django.http import JsonResponse
        try:
            data = _json.loads(request.body)
            order = data.get('order', [])
            for i, task_id in enumerate(order, start=1):
                ServiceTask.objects.filter(pk=int(task_id), route_id=pk).update(route_order=i)
            return JsonResponse({'ok': True})
        except Exception as exc:
            return JsonResponse({'ok': False, 'error': str(exc)}, status=400)

    def add_stop_view(self, request, pk: int):
        if request.method != 'POST':
            from django.http import JsonResponse
            return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
        from django.http import JsonResponse
        try:
            data = _json.loads(request.body)
            task_id = int(data['task_id'])
            route = Route.objects.get(pk=pk)
            task = ServiceTask.objects.get(pk=task_id)
            # Auto-assign next order number
            current_max = ServiceTask.objects.filter(route=route).aggregate(
                m=Max('route_order')
            )['m'] or 0
            task.route = route
            task.route_order = current_max + 1
            task.save(update_fields=['route', 'route_order'])
            return JsonResponse({'ok': True, 'route_order': task.route_order})
        except Exception as exc:
            return JsonResponse({'ok': False, 'error': str(exc)}, status=400)


# ──────────────────────────────────────────────────────────────────────────────
# Filtro personalizado: fecha concreta de tarea
# ──────────────────────────────────────────────────────────────────────────────
class ScheduledDateFilter(admin.SimpleListFilter):
    title          = 'Fecha programada'
    parameter_name = 'scheduled_date_exact'

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)
        dates = (
            qs.order_by('scheduled_date')
            .values_list('scheduled_date', flat=True)
            .distinct()
        )
        return [(d.isoformat(), d.strftime('%d/%m/%Y')) for d in dates]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(scheduled_date=self.value())
        return queryset


# ──────────────────────────────────────────────────────────────────────────────
# Filtro personalizado: tareas con asignación pendiente
# ──────────────────────────────────────────────────────────────────────────────
class PendingAssignmentFilter(admin.SimpleListFilter):
    title          = 'Asignación pendiente'
    parameter_name = 'pending'

    def lookups(self, request, model_admin):
        return [
            ('driver',  'Sin conductor'),
            ('vehicle', 'Sin vehículo'),
            ('both',    'Sin conductor ni vehículo'),
        ]

    def queryset(self, request, queryset):
        if self.value() == 'driver':
            return queryset.filter(driver__isnull=True)
        if self.value() == 'vehicle':
            return queryset.filter(vehicle__isnull=True)
        if self.value() == 'both':
            return queryset.filter(driver__isnull=True, vehicle__isnull=True)
        return queryset


# ──────────────────────────────────────────────────────────────────────────────
# ServiceTask — formulario auxiliar para la Admin Action
# ──────────────────────────────────────────────────────────────────────────────
class ReassignDriverForm(forms.Form):
    new_driver = forms.ModelChoiceField(
        queryset=Driver.objects.all(),
        label='Nuevo conductor',
        empty_label='--- Selecciona un conductor ---',
        help_text='La disponibilidad real se valida dinámicamente por fecha de tarea.',
    )


# ──────────────────────────────────────────────────────────────────────────────
# ServiceTask
# ──────────────────────────────────────────────────────────────────────────────
@admin.register(ServiceTask)
class ServiceTaskAdmin(ModuleFilterMixin, admin.ModelAdmin):
    _module_field = 'contract__module'
    change_list_template = 'admin/rutas/servicetask/change_list.html'
    list_display   = (
        'budget_number_display', 'task_type', 'scheduled_date', 'location',
        'driver', 'vehicle', 'suggested_size_badge', 'route_badge', 'contract',
    )
    list_filter    = (
        'task_type',
        ScheduledDateFilter,
        'driver',
        'vehicle__status',
        'location__zone',
        'location__company',
        'contract__status',
        'route',
        PendingAssignmentFilter,
    )
    search_fields       = (
        'driver__name', 'location__name', 'location__company__name',
        'location__town', 'vehicle__license_plate', 'contract__budget_number',
    )
    date_hierarchy      = 'scheduled_date'
    autocomplete_fields = ('driver', 'vehicle', 'location', 'route')
    list_select_related = ('driver', 'vehicle', 'location', 'contract', 'route')
    actions             = ['delete_selected', 'reassign_driver_action']

    def get_urls(self):
        urls = super().get_urls()
        extra = [
            path(
                'exportar-dia/',
                self.admin_site.admin_view(self.export_day_view),
                name='rutas_servicetask_export_day',
            ),
        ]
        return extra + urls

    def export_day_view(self, request):
        from datetime import timedelta

        target_date = timezone.localdate() + timedelta(days=1)
        if request.method == 'POST':
            date_str = request.POST.get('target_date', '').strip()
            try:
                target_date = date.fromisoformat(date_str)
            except ValueError:
                self.message_user(request, 'Fecha no válida.', messages.ERROR)
            else:
                tasks = (
                    ServiceTask.objects.filter(
                        scheduled_date=target_date,
                        contract__module=_current_module(request),
                    )
                    .select_related('location__company', 'contract')
                    .order_by('location__company__name', 'location__name', 'task_type')
                )

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = 'Tareas'
                headers = [
                    'CLIENTE', 'POBLACION', 'ID EXTERNO', 'Nº PRESUPUESTO',
                    'DIRECCION LINEA 2', 'DIRECCION',
                    'LIMPIEZA', 'UNIDADES', 'COMENTARIOS', 'HORA', 'PERSONA DE REF.',
                    'TELÉFONO', 'EMAIL',
                ]
                sheet.append(headers)

                for task in tasks:
                    location = task.location
                    company = location.company if location else None
                    coords = location.coords_cabin if location else ''
                    budget_num = task.contract.budget_number if task.contract_id else ''
                    hour = ''

                    limpieza_value = ''
                    if task.task_type == ServiceTask.TaskType.LIMPIEZA and task.contract:
                        weekdays = sorted(task.contract.cleaning_weekdays or [])
                        frequency = task.contract.cleaning_frequency or len(weekdays)
                        if frequency <= 1:
                            limpieza_value = 'S'
                        else:
                            try:
                                order = weekdays.index(task.scheduled_date.weekday()) + 1
                            except ValueError:
                                order = 1
                            limpieza_value = f'{order}/{frequency}'

                    comments_value = location.comment if location else ''
                    if task.task_type == ServiceTask.TaskType.ENTREGA:
                        comments_value = 'EO'
                    elif task.task_type == ServiceTask.TaskType.RECOGIDA:
                        comments_value = 'RE'

                    contact_email = (location.email if location and location.email else '') or (company.email if company else '')
                    main_row = [
                        company.name if company else '',
                        (location.town or '').upper() if location else '',
                        location.name if location else '',
                        budget_num,
                        location.address if location else '',
                        coords,
                        limpieza_value,
                        location.cabin_count if location else 1,
                        comments_value,
                        hour,
                        location.contact_name if location else '',
                        location.contact_phone if location else '',
                        contact_email,
                    ]
                    sheet.append(main_row)

                    # Fila extra para la entrada de la finca (si existe)
                    if location and location.coords_entrance:
                        entrance_row = list(main_row)
                        entrance_row[5] = location.coords_entrance  # DIRECCION
                        entrance_row[8] = 'Entrada finca'           # COMENTARIOS
                        sheet.append(entrance_row)

                buffer = BytesIO()
                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = (
                    f'attachment; filename="tareas_{target_date.isoformat()}.xlsx"'
                )
                return response

        return TemplateResponse(
            request,
            'admin/rutas/servicetask/export_day.html',
            {
                'title': 'Exportar tareas por día',
                'target_date': target_date,
                'opts': self.model._meta,
            },
        )

    @admin.display(description='Nº presupuesto', ordering='contract__budget_number')
    def budget_number_display(self, obj: ServiceTask) -> str:
        if obj.contract_id and obj.contract.budget_number:
            return obj.contract.budget_number
        return '—'

    @admin.display(description='Tam. sugerido', ordering='suggested_vehicle_size')
    def suggested_size_badge(self, obj: ServiceTask) -> str:
        if obj.suggested_vehicle_size is None:
            return '—'
        return obj.get_suggested_vehicle_size_display()

    @admin.display(description='Ruta', ordering='route__date')
    def route_badge(self, obj: ServiceTask) -> str:
        if not obj.route_id:
            return '—'
        url = reverse('admin:rutas_route_mapa', args=[obj.route_id])
        label = f'#{obj.route_order or "?"}' if obj.route_order else ''
        return format_html(
            '<a href="{}" title="{}">{} 🗺️</a>',
            url,
            str(obj.route),
            label,
        )

    # ------------------------------------------------------------------
    # Admin Action · Reasignación masiva de conductor
    # ------------------------------------------------------------------
    @admin.action(description='Reasignar conductor a las tareas seleccionadas')
    def reassign_driver_action(self, request, queryset):
        if 'apply' in request.POST:
            form = ReassignDriverForm(request.POST)
            if form.is_valid():
                new_driver: Driver = form.cleaned_data['new_driver']
                updated = queryset.update(driver=new_driver)
                self.message_user(
                    request,
                    f"{updated} tarea(s) reasignada(s) correctamente "
                    f"al conductor «{new_driver.name}».",
                    messages.SUCCESS,
                )
                return None
        else:
            form = ReassignDriverForm()

        return TemplateResponse(
            request,
            'admin/rutas/servicetask/reassign_driver.html',
            {
                'title':  'Reasignar conductor',
                'tasks':  queryset,
                'form':   form,
                'action': 'reassign_driver_action',
                'opts':   self.model._meta,
            },
        )
